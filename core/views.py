# D:\final\core\views.py
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from .models import DailyReport
from .forms import UserUpdateForm, ProfileUpdateForm, UserPasswordChangeForm
from .models import Profile
import pytz
from django.db import models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Sum
from django.contrib import messages
import sys
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import jdatetime
import datetime
from .models import Patient, Company # مطمئن شوید Company هم ایمپورت شده
from .forms import PatientForm # CompanyForm هم باید ایمپورت شود
from django.http import JsonResponse
from django.db import transaction
from django.forms import inlineformset_factory
from django.db.models import Max
from django.core.paginator import Paginator
from django.db.models import Sum, F
from rest_framework import generics, status
from rest_framework.views import APIView # اضافه شده
from rest_framework.response import Response # اضافه شده
from rest_framework.permissions import IsAuthenticated, AllowAny # AllowAny اضافه شده
# از serializers موجود شما استفاده میکنیم و PatientAuthSerializer رو ایمپورت میکنیم
from .serializers import PatientSerializer, PatientAuthSerializer # DrugSerializer اگر لازم بود
from openpyxl import load_workbook
from django_filters.rest_framework import DjangoFilterBackend 
from clinic_messages.models import Message, MessageRecipient 
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login as auth_login
from drugs.models import Drug
from django.contrib.messages.views import SuccessMessageMixin
from core.filters import PatientFilter
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from visits.models import Visit
from drugs.models import  DrugBatch
from django.contrib.auth import get_user_model
User = get_user_model()
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.utils import timezone
from .models import DailyReport, Requirement ,RequirementView
import jdatetime
from django.template.loader import render_to_string
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from fido2.server import Fido2Server
from fido2.webauthn import PublicKeyCredentialRpEntity
import json
import base64
import os
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
@login_required
def passkey_registration_start(request):
    # تولید چالش و شناسه به فرمت استاندارد
    challenge = base64.urlsafe_b64encode(os.urandom(32)).decode('utf-8').replace('=', '')
    user_id = base64.urlsafe_b64encode(str(request.user.id).encode()).decode('utf-8').replace('=', '')
    
    return JsonResponse({
        'challenge': challenge,
        'user': {
            'id': user_id,
            'name': request.user.username,
            'displayName': request.user.get_full_name() or request.user.username,
        }
    })

@csrf_exempt
@login_required
def passkey_registration_complete(request):
    return JsonResponse({'status': 'ok'})
def biometric_challenge(request):
    """ایجاد چالش برای ارسال به مرورگر"""
    registration_data, state = server.register_begin(
        user={
            'id': str(request.user.id).encode(),
            'name': request.user.username,
            'displayName': request.user.get_full_name(),
        },
        credentials=[] # کلیدهای قبلی کاربر از دیتابیس بارگذاری شود
    )
    request.session['webauthn_state'] = state
    return JsonResponse(dict(registration_data))
class CustomLoginView(SuccessMessageMixin, LoginView):
    template_name = 'core/login.html'
    authentication_form = AuthenticationForm
    # redirect_authenticated_user = True # این خط حذف یا کامنت شود
    # success_url = reverse_lazy('dashboard') # این خط حذف یا کامنت شود
    success_message = "با موفقیت وارد شدید!"

    def get_success_url(self):
        user = self.request.user
        if user.is_authenticated:
            # مثال: ریدایرکت بر اساس نقش (گروه)
            if user.groups.filter(name='Doctor').exists() or user.groups.filter(name='Supervisor').exists():
                return reverse_lazy('dashboard') # یا یک داشبورد کلی
            elif user.groups.filter(name='Nurse').exists():
                return reverse_lazy('visits:visit_list') # مثال برای پرستار
            elif user.groups.filter(name='Personnel').exists():
                return reverse_lazy('clinic_messages:message_list') # مثال برای کارگزین (پیام ها)
            elif user.groups.filter(name='Accountant').exists():
                # فرض بر این است که یک URL برای فاکتورها/مالی دارید
                return reverse_lazy('financial:invoice_list') # این را با URL واقعی خود جایگزین کنید
            elif user.groups.filter(name='Supplier').exists():
                # فرض بر این است که یک URL برای فاکتورهای خرید/تامین دارید
                return reverse_lazy('supply:purchase_invoice_list') # این را با URL واقعی خود جایگزین کنید
            else:
                return reverse_lazy('dashboard') # ریدایرکت پیش‌فرض برای نقش‌های نامشخص

        return reverse_lazy('login') # در صورت عدم احراز هویت، به صفحه ورود بازگردانده شود

    def form_invalid(self, form):
        messages.error(self.request, "نام کاربری یا رمز عبور اشتباه است.")
        return super().form_invalid(form)

class CustomLogoutView(LogoutView):
    next_page = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        messages.info(request, "با موفقیت خارج شدید!")
        return super().dispatch(request, *args, **kwargs)



@login_required
def dashboard(request):
    shift_mode = request.GET.get('shift_mode', '12')
    
    # --- اصلاح اساسی برای حل مشکل اختلاف ساعت ---
    now_utc = timezone.now()
    now = timezone.localtime(now_utc) 
    
    today_7am = now.replace(hour=7, minute=0, second=0, microsecond=0)
    
    if shift_mode == '24':
        if now.hour >= 7:
            shift_start = today_7am
            shift_end = shift_start + timedelta(days=1)
        else:
            shift_start = today_7am - timedelta(days=1)
            shift_end = today_7am
        page_title = "داشبورد (شیفت ۲۴ ساعته)"
        current_shift_label = "۲۴ ساعته"
    else:
        if 7 <= now.hour < 19:
            shift_start = today_7am
            shift_end = now.replace(hour=19, minute=0, second=0, microsecond=0)
            page_title = "داشبورد (شیفت صبح)"
            current_shift_label = "شیفت صبح (۷ تا ۱۹)"
        else:
            if now.hour >= 19:
                shift_start = now.replace(hour=19, minute=0, second=0, microsecond=0)
                shift_end = shift_start + timedelta(days=1)
            else:
                shift_start = (now - timedelta(days=1)).replace(hour=19, minute=0, second=0, microsecond=0)
                shift_end = today_7am
            page_title = "داشبورد (شیفت شب)"
            current_shift_label = "شیفت شب (۱۹ تا ۷)"

    welcome_message = f"خوش آمدید، {request.user.get_full_name() or request.user.username}"
    
    j_now = jdatetime.datetime.now()
    j_first_day = jdatetime.datetime(j_now.year, j_now.month, 1)
    g_first_day = j_first_day.togregorian()
    current_month_name = j_now.strftime('%B')

    shift_visits_query = Visit.objects.filter(visit_date__range=[shift_start, shift_end])\
                                      .select_related('patient', 'assigned_to')\
                                      .order_by('-visit_date')

    reasons_shift = Visit.objects.filter(visit_date__range=[shift_start, shift_end])\
        .values('reason_for_visit__name')\
        .annotate(count=Count('id')).order_by('-count')

    reasons_month = Visit.objects.filter(visit_date__gte=g_first_day)\
        .values('reason_for_visit__name')\
        .annotate(count=Count('id')).order_by('-count')

    results_shift = Visit.objects.filter(visit_date__range=[shift_start, shift_end])\
        .values('treatment_result__name')\
        .annotate(count=Count('id')).order_by('-count')

    results_month = Visit.objects.filter(visit_date__gte=g_first_day)\
        .values('treatment_result__name')\
        .annotate(count=Count('id')).order_by('-count')

    total_month_visits = Visit.objects.filter(visit_date__gte=g_first_day).count()
    
    user_stats = User.objects.annotate(
        visit_count=Count('visits_as_doctor', filter=Q(visits_as_doctor__visit_date__gte=g_first_day))
    ).filter(visit_count__gt=0).order_by('-visit_count')

    for u in user_stats:
        if total_month_visits > 0:
            u.share_percent = round((u.visit_count / total_month_visits) * 100, 1)
        else:
            u.share_percent = 0

    expiring_drugs = DrugBatch.objects.filter(
        expiry_date__range=[now.date(), now.date() + timedelta(days=60)],
        quantity__gt=0
    ).order_by('expiry_date')

    # --- بخش اضافه شده برای هشدار موجودی کم ---
    low_stock_drugs = Drug.objects.annotate(
        total_qty=Coalesce(Sum('batches__quantity'), 0)
    ).filter(
        total_qty__lte=F('min_stock_alert')
    ).order_by('total_qty')

    user_groups = request.user.groups.values_list('name', flat=True)
    is_medical_staff = any(role in user_groups for role in ['Doctor', 'Supervisor', 'Nurse'])

    context = {
        'page_title': page_title,
        'welcome_message': welcome_message,
        'current_month_name': current_month_name,
        'shift_mode': shift_mode,
        'shift_start': shift_start,
        'shift_end': shift_end,
        'current_shift_label': current_shift_label,
        'total_visits_shift': shift_visits_query.count(),
        'shift_visits': shift_visits_query,
        'reasons_shift': reasons_shift,
        'reasons_month': reasons_month,
        'results_shift': results_shift,
        'results_month': results_month,
        'user_stats': user_stats,
        'total_month_visits': total_month_visits,
        'expiring_drugs': expiring_drugs,
        'low_stock_drugs': low_stock_drugs,
        'is_medical_staff': is_medical_staff,
        'is_doctor': 'Doctor' in user_groups,
        'is_nurse': 'Nurse' in user_groups,
        'is_supervisor': 'Supervisor' in user_groups,
    }
    return render(request, 'core/dashboard.html', context)
def patient_list(request):
    base_queryset = Patient.objects.select_related('company').all().order_by('-pk')
    
    # گرفتن مقدار جستجو از لیست (حتی اگر چند تا q ارسال شده باشد، اولین مقدار غیرخالی را می‌گیرد)
    queries = request.GET.getlist('q')
    query = next((item for item in queries if item), '').strip()
    
    # فیلتر کردن کوئری‌ست اصلی
    base_queryset = base_queryset.filter(
        Q(first_name__icontains=query) | 
        Q(last_name__icontains=query) | 
        Q(national_code__icontains=query) |
        Q(phone_number__icontains=query) |  # اضافه شد
        Q(personnel_number__icontains=query) # اضافه شد
    ).distinct()

    patient_filter = PatientFilter(request.GET, queryset=base_queryset)
    paginator = Paginator(patient_filter.qs, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_title': 'لیست بیماران',
        'page_obj': page_obj,
        'filter': patient_filter,
        'search_query': query,
    }

    # تشخیص درخواست AJAX برای آپدیت بخشی از صفحه
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('core/patient_list_partial.html', context, request=request)
        return JsonResponse({'html': html})

    return render(request, 'core/patient_list.html', context)
    
@login_required
def patient_create(request):
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES)
        if form.is_valid():
            patient = form.save(commit=False)
            patient.registered_by = request.user
            patient.save()
            
            # ⭐ این بخش اصلاح شده است تا بر اساس دکمه کلیک شده، عمل کند. ⭐
            if 'save_and_visit' in request.POST:
                messages.success(request, 'بیمار با موفقیت ثبت شد. لطفاً ویزیت جدید را ثبت کنید.')
                # هدایت به صفحه ثبت ویزیت با شناسه بیمار جدید
                return redirect('visits:visit_create_for_patient', patient_id=patient.id)
            else:
                messages.success(request, 'بیمار با موفقیت ثبت شد.')
                # هدایت پیش‌فرض به لیست بیماران
                return redirect('core:patient_list')
        else:
            messages.error(request, 'لطفا خطاهای فرم را برطرف کنید.')
    else:
        form = PatientForm()
    
    context = {
        'form': form,
        'page_title': 'ثبت بیمار جدید'
    }
    return render(request, 'core/patient_form.html', context)

@login_required
def patient_update(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES, instance=patient)
        if form.is_valid():
            patient = form.save()
            messages.success(request, 'اطلاعات بیمار با موفقیت به‌روزرسانی شد.')
            return redirect('core:patient_list')
        else:
            messages.error(request, 'لطفا خطاهای فرم را برطرف کنید.')
    else:
        form = PatientForm(instance=patient)
        
    context = {
        'form': form,
        'page_title': f'ویرایش بیمار: {patient.full_name}'
    }
    return render(request, 'core/patient_form.html', context)

@login_required(login_url=reverse_lazy('login'))
def patient_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    # مطمئن شوید مدل Visit وجود دارد و به Patient مرتبط است
    # visits = patient.visits.all().order_by('-visit_date') 
    # فعلاً به خاطر عدم وجود مدل Visit در اینجا، این خط رو کامنت میکنیم.
    visits = [] # placeholder
    context = {
        'page_title': f'جزئیات بیمار: {patient.full_name}',
        'patient': patient,
        'visits': visits
    }
    return render(request, 'core/patient_detail.html', context)

@login_required(login_url=reverse_lazy('login'))
@permission_required('core.delete_patient', raise_exception=True)
def patient_delete(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        patient.delete()
        messages.success(request, 'بیمار با موفقیت حذف شد!')
        return redirect('core:patient_list')
    context = {
        'page_title': 'حذف بیمار',
        'object': patient
    }
    return render(request, 'core/patient_list.html', context)

@login_required(login_url=reverse_lazy('login'))
def company_list(request):
    companies = Company.objects.all().order_by('name')
    context = {
        'page_title': 'لیست شرکت‌ها',
        'companies': companies
    }
    return render(request, 'core/company_list.html', context)

@login_required(login_url=reverse_lazy('login'))
def company_create(request):
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'شرکت با موفقیت اضافه شد!')
            return redirect('company_list')
        else:
            messages.error(request, 'خطا در ثبت شرکت. لطفا اطلاعات وارد شده را بررسی کنید.')
    else:
        form = CompanyForm()
    context = {
        'page_title': 'افزودن شرکت جدید',
        'form': form
    }
    return render(request, 'core/company_form.html', context)

@login_required(login_url=reverse_lazy('login'))
def company_detail(request, pk):
    company = get_object_or_404(Company, pk=pk)
    context = {
        'page_title': f'جزئیات شرکت: {company.name}',
        'company': company
    }
    return render(request, 'core/company_detail.html', context)

@login_required(login_url=reverse_lazy('login'))
def company_update(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, 'اطلاعات شرکت با موفقیت ویرایش شد!')
            return redirect('company_list')
        else:
            messages.error(request, 'خطا در ویرایش شرکت. لطفا اطلاعات وارد شده را بررسی کنید.')
    else:
        form = CompanyForm(instance=company)
    context = {
        'page_title': f'ویرایش شرکت: {company.name}',
        'form': form
    }
    return render(request, 'core/company_form.html', context)

@login_required(login_url=reverse_lazy('login'))
@permission_required('core.delete_company', raise_exception=True)
def company_delete(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        company.delete()
        messages.success(request, 'شرکت با موفقیت حذف شد!')
        return redirect('company_list')
    context = {
        'page_title': 'حذف شرکت',
        'object': company
    }
    return render(request, 'core/confirm_delete.html', context)

@login_required
def user_profile(request):
    return render(request, 'core/user_profile.html', {'user': request.user})


# --- API Views for React (جدید) ---

class RegisterPatientAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        # 👈👈👈 context={'request': request} رو به سریالایزر پاس بده
        serializer = PatientAuthSerializer(data=request.data, context={'request': request}) 
        if serializer.is_valid():
            try:
                patient = serializer.save() 
                return Response({
                    "message": "درخواست ثبت نام با موفقیت ارسال شد و در انتظار تایید مدیر سیستم است.", # 👈 پیام رو دقیق تر کن
                    "patient_id": patient.id,
                    "user": PatientAuthSerializer(patient, context={'request': request}).data # ارسال داده‌های بیمار ثبت نام شده
                }, status=status.HTTP_201_CREATED)
            except Exception as e:
                print(f"Error during patient registration: {e}") 
                return Response({"detail": "خطای سرور هنگام ثبت نام: " + str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        print(f"Serializer errors: {serializer.errors}") 
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class LoginPatientAPIView(APIView):
    permission_classes = [AllowAny] # برای لاگین، AllowAny مناسب است

    def post(self, request):
        personnel_number = request.data.get('personnelNumber') 
        national_code_input = request.data.get('nationalCode') 

        if not personnel_number or not national_code_input:
            return Response({"detail": "کد پرسنلی و کد ملی (رمز عبور) الزامی است."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            patient = Patient.objects.get(personnel_number=personnel_number)
            
            # --- 👈👈👈 این بخش جدید و حیاتی برای بررسی وضعیت تایید است ---
            if not patient.is_approved:
                return Response({
                    "detail": "حساب کاربری شما هنوز توسط مدیر تایید نشده است. لطفا منتظر بمانید."
                }, status=status.HTTP_403_FORBIDDEN) # 403 Forbidden برای دسترسی ممنوع
            # --- پایان بخش جدید ---

            # --- هشدار امنیتی مهم ---
            # مقایسه مستقیم national_code_input با patient.national_code کاملاً ناامن است.
            # رمز عبور (national_code) باید در دیتابیس هش شده ذخیره شود و هنگام ورود، ورودی کاربر هم هش شده و مقایسه شود.
            # برای مثال، اگر از Django's default User model استفاده می‌کنید، از patient.check_password(national_code_input) استفاده می‌کنید.
            # اگر national_code را هش نکرده‌اید، این سیستم به شدت آسیب‌پذیر است.
            # TODO: در آینده، برای امنیت بیشتر، شماره ملی (که نقش رمز عبور دارد) باید هش شود.
            if patient.national_code == national_code_input: # فرض بر این است که national_code هش نشده است.
                serializer = PatientAuthSerializer(patient)
                
                # --- افزودن توکن به پاسخ (مثال با Simple JWT) ---
                # اگر از Simple JWT استفاده می‌کنید و patient یک User model باشد:
                # refresh = RefreshToken.for_user(patient) 
                # token_data = {
                #     'refresh': str(refresh),
                #     'access': str(refresh.access_token),
                # }
                # return Response({"message": "ورود موفقیت آمیز", "user": serializer.data, "tokens": token_data}, status=status.HTTP_200_OK)

                # در غیر این صورت، پاسخ ساده فعلی:
                return Response({"message": "ورود موفقیت آمیز", "user": serializer.data}, status=status.HTTP_200_OK)
            else:
                return Response({"detail": "کد ملی (رمز عبور) اشتباه است."}, status=status.HTTP_401_UNAUTHORIZED)
        except Patient.DoesNotExist:
            return Response({"detail": "کاربری با این کد پرسنلی یافت نشد."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Error during patient login: {e}") 
            return Response({"detail": "خطای سرور هنگام ورود: " + str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# View برای مدیریت EmergencyAlert ها (در آینده)
class EmergencyAlertAPIView(APIView):
    permission_classes = [IsAuthenticated] # فقط کاربران لاگین شده میتوانند هشدار ارسال کنند

    def post(self, request):
        # از کاربر لاگین شده (که فرض میکنیم یک Patient است) به عنوان فرستنده استفاده میکنیم
        # TODO: این بخش نیاز به تعریف مدل EmergencyAlert و Serializer مربوطه دارد
        # فرض میکنیم request.user به یک Patient متصل است یا میتوانیم از personnel_number ارسالی استفاده کنیم

        # اطلاعاتی که از فرانت اند می آید:
        # {
        #   "user_personnel_id": "P1001",
        #   "location": { "lat": 35.xxx, "lng": 51.xxx, "address": "..." },
        #   "timestamp": "YYYY-MM-DD HH:MM:SS",
        #   "type": "حادثه اضطراری"
        # }
        
        user_personnel_id = request.data.get('user_personnel_id')
        location_data = request.data.get('location')
        timestamp = request.data.get('timestamp')
        incident_type = request.data.get('type')

        if not user_personnel_id or not location_data:
            return Response({"detail": "اطلاعات کاربر و موقعیت مکانی الزامی است."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            patient = Patient.objects.get(personnel_number=user_personnel_id)
            # TODO: اینجا باید یک شیء EmergencyAlert ایجاد و ذخیره شود
            # EmergencyAlert.objects.create(
            #     patient=patient,
            #     latitude=location_data.get('lat'),
            #     longitude=location_data.get('lng'),
            #     address_description=location_data.get('address'),
            #     incident_type=incident_type,
            #     timestamp=timestamp # اگر timestamp را به عنوان string میفرستیم، باید parse شود
            # )
            
            print(f"Emergency Alert Received for {patient.full_name}: Type={incident_type}, Location={location_data}, Time={timestamp}")
            return Response({"message": "گزارش حادثه با موفقیت دریافت شد."}, status=status.HTTP_200_OK)
        except Patient.DoesNotExist:
            return Response({"detail": "بیمار با کد پرسنلی ارسال شده یافت نشد."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# --------------------------------------------------
# پروفایل کاربر (User Profile) - این ویو HTML برمی‌گرداند، نه JSON
# --------------------------------------------------
@login_required
def user_profile(request):
    return render(request, 'core/user_profile.html', {'user': request.user})


# لیست Patient ها برای API (اگر نیاز باشد)
# این همان generics.ListAPIView است که در کامنت قبلی داشتید
class PatientListAPIView(generics.ListAPIView):
    queryset = Patient.objects.all().order_by('last_name', 'first_name')
    serializer_class = PatientSerializer # از PatientSerializer موجود شما استفاده میکند
    filter_backends = [DjangoFilterBackend]
    filterset_class = PatientFilter # اگر فیلتر برای API نیاز دارید
    permission_classes = [IsAuthenticated] # فقط کاربران احراز هویت شده می‌توانند این لیست را ببینند
 
def upload_personnel_file(request):
    """
    نمایش صفحه آپلود فایل اکسل مشخصات پرسنل.
    """
    return render(request, 'core/upload_personnel.html')


def process_personnel_import(request):
    """
    پردازش فایل اکسل مشخصات پرسنل:
    - اگر کد ملی وجود داشته باشد، شماره پرسنلی، نام، نام خانوادگی، و شغل را به‌روزرسانی می‌کند.
    - اگر کد ملی وجود نداشته باشد، بیمار جدیدی را ثبت می‌کند.
    - شرکت برای بیماران موجود در فایل اکسل، به شرکت مشخص شده در فرم، به‌روزرسانی می‌شود.
    """
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'لطفاً یک فایل اکسل انتخاب کنید.')
            return redirect('core:upload_personnel_file')

        excel_file = request.FILES['excel_file']

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'فایل انتخاب شده باید از نوع اکسل (xlsx یا xls) باشد.')
            return redirect('core:upload_personnel_file')

        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            # خواندن و پاک کردن فاصله‌های خالی از سربرگ‌ها
            header = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]

            # تعریف نگاشت سربرگ‌های اکسل به فیلدهای مدل Patient
            column_mapping = {
                'نام': 'first_name',
                'نام خانوادگی': 'last_name',
                'کد ملی': 'national_code',
                'شماره پرسنلی': 'personnel_number',
                'واحد سازمانی 2': 'occupation', # نگاشت ستون اکسل 'واحد سازمانی 2' به فیلد 'occupation' در مدل Patient
            }

            # پیدا کردن ایندکس ستون‌ها بر اساس سربرگ
            col_indices = {}
            for excel_header, model_field in column_mapping.items():
                try:
                    col_indices[model_field] = header.index(excel_header)
                except ValueError:
                    messages.warning(request, f'سربرگ "{excel_header}" در فایل اکسل یافت نشد. این ستون نادیده گرفته خواهد شد.')
                    col_indices[model_field] = -1

            if col_indices.get('national_code', -1) == -1:
                messages.error(request, 'ستون "کد ملی" در فایل اکسل ضروری است و یافت نشد.')
                return redirect('core:upload_personnel_file')

            # --- تغییرات اضافه شده برای مدیریت شرکت ---
            company_name_from_user = request.POST.get('company_name', '').strip()
            if not company_name_from_user:
                messages.error(request, 'نام شرکت الزامی است.')
                return redirect('core:upload_personnel_file')
            
            try:
                company_obj, created_company = Company.objects.get_or_create(name=company_name_from_user)
                if created_company:
                    messages.info(request, f"شرکت '{company_name_from_user}' ایجاد شد.")
            except Exception as e:
                messages.error(request, f"خطا در پیدا کردن یا ایجاد شرکت '{company_name_from_user}': {e}")
                return redirect('core:upload_personnel_file')
            # --- پایان تغییرات مدیریت شرکت ---

            updated_count = 0
            created_count = 0
            skipped_count = 0

            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    row_data = [cell.value for cell in row]

                    national_code = None
                    if col_indices['national_code'] != -1:
                        national_code = str(row_data[col_indices['national_code']]).strip() if row_data[col_indices['national_code']] else None

                    if not national_code:
                        messages.warning(request, f'ردیف {row_idx}: کد ملی خالی است. این ردیف نادیده گرفته شد.')
                        skipped_count += 1
                        continue

                    personnel_number = None
                    if col_indices['personnel_number'] != -1:
                        personnel_number = str(row_data[col_indices['personnel_number']]).strip() if row_data[col_indices['personnel_number']] else None

                    first_name = None
                    if col_indices['first_name'] != -1:
                        first_name = str(row_data[col_indices['first_name']]).strip() if row_data[col_indices['first_name']] else ''

                    last_name = None
                    if col_indices['last_name'] != -1:
                        last_name = str(row_data[col_indices['last_name']]).strip() if row_data[col_indices['last_name']] else ''
                    
                    # خواندن مقدار 'occupation' از اکسل
                    occupation = None
                    if col_indices['occupation'] != -1:
                        occupation = str(row_data[col_indices['occupation']]).strip() if row_data[col_indices['occupation']] else ''

                    try:
                        # تلاش برای دریافت یا ایجاد بیمار بر اساس کد ملی
                        patient, created = Patient.objects.get_or_create(
                            national_code=national_code,
                            defaults={
                                'first_name': first_name,
                                'last_name': last_name,
                                'personnel_number': personnel_number,
                                'occupation': occupation, # مقدار occupation از اکسل
                                'company': company_obj, # شرکت تنظیم شده از ورودی کاربر
                            }
                        )
                        if created:
                            created_count += 1
                            messages.success(request, f'ردیف {row_idx}: بیمار جدید با کد ملی {national_code} و شماره پرسنلی {personnel_number} ثبت شد.')
                        else:
                            # اگر بیمار از قبل موجود بود، فیلدها را در صورت نیاز به‌روزرسانی می‌کنیم
                            needs_update = False
                            if personnel_number is not None and patient.personnel_number != personnel_number:
                                patient.personnel_number = personnel_number
                                needs_update = True
                            
                            if first_name and patient.first_name != first_name:
                                patient.first_name = first_name
                                needs_update = True

                            if last_name and patient.last_name != last_name:
                                patient.last_name = last_name
                                needs_update = True
                            
                            # به‌روزرسانی فیلد شغل اگر مقدار جدیدی از اکسل داریم و متفاوت است
                            if occupation is not None and patient.occupation != occupation:
                                patient.occupation = occupation
                                needs_update = True

                            # اطمینان از اینکه شرکت همیشه به مقدار ورودی کاربر تنظیم شود (فقط برای این بیمار)
                            if patient.company != company_obj:
                                patient.company = company_obj
                                needs_update = True

                            if needs_update:
                                patient.save()
                                updated_count += 1
                                messages.info(request, f'ردیف {row_idx}: اطلاعات بیمار با کد ملی {national_code} به‌روزرسانی شد.')
                            else:
                                messages.info(request, f'ردیف {row_idx}: بیمار با کد ملی {national_code} از قبل موجود بود و نیازی به به‌روزرسانی نبود.')

                    except Exception as e:
                        messages.error(request, f'ردیف {row_idx}: خطایی در پردازش رخ داد: {e}')
                        skipped_count += 1
                
                # --- این خط کد قبلاً باعث به‌روزرسانی همه بیماران می‌شد و اکنون حذف شده است ---
                # bulk_updated_patients_count = Patient.objects.update(company=company_obj)
                # messages.info(request, f'تعداد {bulk_updated_patients_count} بیمار به شرکت "{company_name_from_user}" منتقل شدند.')
                # --- پایان حذف ---

            messages.success(request, f'فرآیند آپلود پرسنل با موفقیت به پایان رسید. تعداد ثبت شده: {created_count}، تعداد به‌روزرسانی شده: {updated_count}، تعداد نادیده گرفته شده: {skipped_count}.')

        except Exception as e:
            messages.error(request, f'خطا در خواندن فایل اکسل: {e}')
            return redirect('core:upload_personnel_file')

    return redirect('core:upload_personnel_file')

@login_required
def user_profile(request):

    # چک کنید که آیا مدل پروفایل برای کاربر وجود دارد یا نه
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        profile = Profile.objects.create(user=request.user)
    
    # برای مدیریت درخواست های POST
    if request.method == 'POST':
        # چک کنید کدام فرم ارسال شده است
        if 'update_user' in request.POST:
            user_form = UserUpdateForm(request.POST, instance=request.user)
            profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
            if user_form.is_valid() and profile_form.is_valid():
                user_form.save()
                profile_form.save()
                messages.success(request, 'اطلاعات پروفایل با موفقیت به‌روزرسانی شد.')
                return redirect('core:user_profile')
        
        elif 'change_password' in request.POST:
            password_form = UserPasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user) # مهم: برای جلوگیری از لاگ اوت شدن کاربر
                messages.success(request, 'رمز عبور با موفقیت تغییر یافت.')
                return redirect('core:user_profile')
            else:
                messages.error(request, 'خطا در تغییر رمز عبور. لطفا اطلاعات را مجددا بررسی کنید.')
        
        # اگر فرم ها نامعتبر بودند
        else:
            user_form = UserUpdateForm(instance=request.user)
            profile_form = ProfileUpdateForm(instance=profile)
            password_form = UserPasswordChangeForm(request.user)
            messages.error(request, 'خطا در ارسال فرم.')

    # برای مدیریت درخواست های GET
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=profile)
        password_form = UserPasswordChangeForm(request.user)

    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'password_form': password_form,
        'page_title': 'پروفایل کاربری'
    }
    return render(request, 'core/profile.html', context)




def get_exact_shift_times(raw_date, shift_value):
    tehran_tz = pytz.timezone('Asia/Tehran')
    
    # تنظیم زمان به وقت تهران
    if timezone.is_naive(raw_date):
        now_dt = timezone.make_aware(raw_date, tehran_tz)
    else:
        now_dt = raw_date.astimezone(tehran_tz)

    # اگر قبل از ۷:۱۰ صبح است، متعلق به شیفت روز قبل است
    if now_dt.hour < 7 or (now_dt.hour == 7 and now_dt.minute < 10):
        operational_date = now_dt - datetime.timedelta(days=1)
    else:
        operational_date = now_dt

    # نقطه شروع مبنا: ۷:۱۰ صبح روز عملیاتی
    start_7_10_am = operational_date.replace(hour=7, minute=10, second=0, microsecond=0)

    # --- شروع منطق تفکیک شیفت‌ها ---
    
    # ۱. شیفت ۱۲ ساعته صبح (۷:۱۰ تا ۱۹:۱۰)
    if 'صبح' in shift_value or shift_value == '12_morning':
        start = start_7_10_am
        end = start_7_10_am.replace(hour=19, minute=10, second=0)
        label = "۱۲ ساعته صبح"

    # ۲. شیفت ۱۲ ساعته شب (۱۹:۱۰ تا ۷:۱۰ فردا)
    elif 'شب' in shift_value or shift_value == '12_night':
        start = start_7_10_am.replace(hour=19, minute=10, second=0)
        end = start + datetime.timedelta(hours=12)
        label = "۱۲ ساعته شب"

    # ۳. شیفت ۲۴ ساعته (۷:۱۰ امروز تا ۷:۱۰ فردا)
    else:
        start = start_7_10_am
        end = start + datetime.timedelta(hours=24)
        label = "۲۴ ساعته"
    
    return start, end, label
@login_required
def manage_daily_report(request):
    if request.method == "POST":
        shift_val = request.POST.get('shift_selection')
        now = timezone.now()
        start, end, label = get_exact_shift_times(now, shift_val)

        report = DailyReport.objects.create(
            doctor_id=request.POST.get('doctor'),
            nurse=request.user,
            driver_id=request.POST.get('driver'),
            shift_type=label, # ذخیره لیبل فارسی (صبح/شب)
            # سایر فیلدها...
            dispatched_action=request.POST.get('dispatched_action', ''),
            next_shift_plan=request.POST.get('next_shift_plan', ''),
            ambulance_status=request.POST.get('ambulance_status', ''),
            equipment_status=request.POST.get('equipment_status', ''),
        )
        
        # مدیریت نیازمندی‌ها
        new_reqs = request.POST.getlist('new_requirements[]')
        for text in new_reqs:
            if text.strip():
                nr = Requirement.objects.create(title=text.strip(), creator=request.user)
                report.requirements.add(nr)
        
        return redirect('core:print_daily_report', report_id=report.id)

    # بخش GET
    context = {
        'doctors': User.objects.filter(groups__name='doctor'),
        'drivers': User.objects.filter(groups__name__in=['Nurse', 'driver']),
        'j_date': jdatetime.datetime.now().strftime('%Y/%m/%d'),
    }
    return render(request, 'core/daily_report_form.html', context)

@login_required
def print_daily_report(request, report_id):
    report = get_object_or_404(DailyReport, id=report_id)
    
    # محاسبه زمان شیفت (با همان منطق 7:10 که خواستی)
    # اگر فیلد created_at خالی بود از date استفاده کن
    report_time = report.created_at or timezone.now()
    start, end, label = get_exact_shift_times(report_time, report.shift_type)

    visits = Visit.objects.filter(visit_date__range=[start, end])

    stats = {
        'total': visits.count(),
        'dispatched_count': visits.filter(treatment_result__name__icontains='اعزام').count(),
        'reasons': [f"{i['reason_for_visit__name']}: {i['count']} مورد" for i in visits.values('reason_for_visit__name').annotate(count=Count('id')).order_by('-count')[:6] if i['reason_for_visit__name']],
        'results': [f"{i['treatment_result__name']}: {i['count']} مورد" for i in visits.values('treatment_result__name').annotate(count=Count('id')).order_by('-count')[:7] if i['treatment_result__name']],
    }

    # تبدیل تاریخ به شمسی برای هدر
    j_date_display = jdatetime.date.fromgregorian(date=start.date()).strftime("%Y/%m/%d")

    context = {
        'report': report, # با فرستادن خودِ report، تمام فیلدهای مدل در HTML قابل دسترسی هستند
        'stats': stats,
        'j_date': j_date_display,
    }
    return render(request, 'core/daily_report_print.html', context)

   
@login_required
def daily_report_list(request):
    # دریافت همه گزارش‌ها به ترتیب جدیدترین
    reports_list = DailyReport.objects.all().order_by('-id')
    
    # فیلتر جستجو (اختیاری)
    search_query = request.GET.get('search')
    if search_query:
        reports_list = reports_list.filter(
            Q(doctor__last_name__icontains=search_query) | 
            Q(nurse__last_name__icontains=search_query)
        )

    context = {
        'reports': reports_list,
    }
    return render(request, 'core/daily_report_list.html', context)

@login_required
def requirement_tracking(request):
    report_id = request.GET.get('report_id')
    
    if report_id:
        report = get_object_or_404(DailyReport, id=report_id)
        requirements = report.requirements.all().order_by('-id')
        # تبدیل تاریخ به شمسی برای عنوان صفحه
        j_date = jdatetime.date.fromgregorian(date=report.created_at.date()).strftime('%Y/%m/%d')
        title_prefix = f"نیازمندی‌های گزارش مورخ {j_date}"
    else:
        requirements = Requirement.objects.all().order_by('-id')
        title_prefix = "لیست کل نیازمندی‌های ثبت شده"

    # --- منطق جدید: ثبت بازدید مدیران ---
    if request.user.is_superuser or request.user.is_staff:
        for req in requirements:
            # اگر مدیر خودش سازنده نباشد، بازدیدش ثبت شود
            if req.creator != request.user:
                # ثبت در جدول واسط (اگر قبلاً ثبت نشده باشد)
                RequirementView.objects.get_or_create(
                    requirement=req, 
                    admin=request.user
                )
                # آپدیت وضعیت به مشاهده شده (اگر هنوز در حالت ایجاد شده باشد)
                if req.status == 'created':
                    req.status = 'viewed'
                    req.save()

    return render(request, 'core/requirement_tracking.html', {
        'requirements': requirements,
        'title_prefix': title_prefix
    })

def mark_as_viewed_logic(current_user, requirements):
    """ثبت بازدید مدیران - اصلاح شده برای رفع AttributeError"""
    if current_user.is_superuser or current_user.is_staff:
        for req in requirements:
            # بررسی اینکه سازنده خودِ مدیر نباشد
            if req.creator != current_user:
                # ثبت بازدید در مدل واسط (get_or_create خودش چک می‌کند که تکراری نباشد)
                RequirementView.objects.get_or_create(
                    requirement=req, 
                    admin=current_user
                )
                # اگر وضعیت 'ایجاد شده' بود، به 'مشاهده شده' تغییر کند
                if req.status == 'created':
                    req.status = 'viewed'
                    req.save()

@login_required
@login_required
def all_requirements_report(request):
    """لیست جامع تمام نیازمندی‌ها"""
    # دریافت نیازمندی‌ها
    requirements = Requirement.objects.all().order_by('-created_at')
    
    # فیلترها
    status_filter = request.GET.get('status')
    if status_filter:
        requirements = requirements.filter(status=status_filter)
        
    creator_filter = request.GET.get('creator')
    if creator_filter:
        requirements = requirements.filter(creator__last_name__icontains=creator_filter)

    # --- اجرای منطق مشاهده مدیران ---
    # پاس دادن مستقیم request.user
    mark_as_viewed_logic(request.user, requirements)

    return render(request, 'core/all_requirements_list.html', {
        'requirements': requirements
    })

@login_required
@require_POST
def update_requirement_status(request, req_id):
    """بروزرسانی وضعیت و یادداشت توسط مدیر"""
    req = get_object_or_404(Requirement, id=req_id)
    req.status = request.POST.get('status')
    req.admin_note = request.POST.get('admin_note')
    
    if req.status == 'resolved':
        req.is_archived = True
        
    req.save()
    return JsonResponse({'status': 'success'})


