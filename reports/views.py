import datetime
from django.shortcuts import render
from django.db.models import Value, CharField, F, fields ,Sum
from django.db import models
from persiantools.jdatetime import JalaliDate, JalaliDateTime
from django.shortcuts import render
from django.db.models import Value, CharField, F, fields
from django.utils import timezone  # حیاتی برای حل مشکل Offset-Aware
from persiantools.jdatetime import JalaliDate, JalaliDateTime
from .filters import DrugTransactionFilter
from django.shortcuts import render
from django.db.models import Count, Q, Sum, F, ExpressionWrapper, fields, OuterRef, Subquery ,Max# OuterRef, Subquery اضافه شدند
from django.db.models.functions import TruncMonth, TruncDay, Coalesce # Coalesce اضافه شد
from django.contrib.auth.decorators import login_required
from datetime import datetime, date, timedelta # timedelta اضافه شد
from persiantools.jdatetime import JalaliDate
import json
from decimal import Decimal # برای BMI و محاسبات دقیق
import collections # برای هیستوگرام فشار خون
from drugs.models import Drug, DrugRequest, DrugRequestItem, PurchaseInvoice, PurchaseInvoiceItem, Supplier # 👈 ایمپورت‌های جدید
from core.models import GENDER_CHOICES, Patient, Company, BLOOD_TYPE_CHOICES
from visits.models import Visit, ReasonForVisit, TreatmentResult, VISIT_STATUS_CHOICES, INCIDENT_TYPE_CHOICES, VisitItem # VisitItem اضافه شد
import re
from reports.filters import PatientFilter, VisitFilter, DrugFilter,DrugTransactionFilter
from drugs.models import DrugBatch 
import jdatetime
from lab_results.models import PeriodicExamination, LabParameterResult 
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import jdatetime
from datetime import timedelta
import io
@login_required
def base_report_context(request, report_title, filter_form=None):
   
    now = datetime.now()
    current_jalali_date = JalaliDate.today().strftime('%Y/%m/%d')
    current_time = now.strftime('%H:%M')
    
    context = {
        'report_title': report_title,
        'filter_form': filter_form, # این فرم فیلتر باید به قالب ارسال شود
        'current_jalali_date': current_jalali_date,
        'current_time': current_time,
        'user': request.user,
    }
    return context





@login_required
def patient_report_view(request):
    """
    گزارش تحلیلی بیماران. داده‌ها تنها پس از اعمال فیلتر بارگذاری و نمایش داده می‌شوند.
    """
    report_title = "گزارش بیماران"
    
    # بررسی اینکه آیا فرم فیلتر سابمیت شده است یا خیر
    # bool(request.GET) چک می‌کند که آیا حداقل یک پارامتر در URL وجود دارد
    is_filtered = bool(request.GET)
    
    # اگر فیلتری اعمال شده، از کل بیماران کوئری بگیر، در غیر این صورت یک کوئری خالی برگردان
    if is_filtered:
        queryset = Patient.objects.select_related('company', 'registered_by').all()
    else:
        queryset = Patient.objects.none()

    patient_filter = PatientFilter(request.GET, queryset=queryset)
    filtered_patients_qs = patient_filter.qs
    
    context = base_report_context(request, report_title, filter_form=patient_filter.form)
    context['is_filtered'] = is_filtered # ارسال این فلگ به تمپلیت

    # تمام محاسبات و آماده‌سازی داده‌ها فقط در صورتی انجام می‌شود که فیلتر اعمال شده باشد
    if is_filtered:
        patients_with_visits = filtered_patients_qs.annotate(
            total_visits=Count('visits')
        ).order_by('-total_visits', 'last_name')

        # توزیع جنسیت
        gender_distribution = filtered_patients_qs.values('gender').annotate(count=Count('id')).order_by('gender')
        gender_chart_labels = [dict(GENDER_CHOICES).get(item['gender'], 'نامشخص') for item in gender_distribution]
        gender_chart_data = [item['count'] for item in gender_distribution]

        # توزیع سنی
        today_year = date.today().year
        age_groups_db = filtered_patients_qs.filter(date_of_birth__isnull=False).annotate(
            age_years=ExpressionWrapper(today_year - F('date_of_birth__year'), output_field=fields.IntegerField())
        ).values('age_years').annotate(count=Count('id')).order_by('age_years')
        age_groups_data = {'0-10': 0, '11-20': 0, '21-30': 0, '31-40': 0, '41-50': 0, '51-60': 0, '>60': 0, 'نامشخص': 0}
        for item in age_groups_db:
            age = item['age_years']
            if 0 <= age <= 10: age_groups_data['0-10'] += item['count']
            elif 11 <= age <= 20: age_groups_data['11-20'] += item['count']
            elif 21 <= age <= 30: age_groups_data['21-30'] += item['count']
            elif 31 <= age <= 40: age_groups_data['31-40'] += item['count']
            elif 41 <= age <= 50: age_groups_data['41-50'] += item['count']
            elif 51 <= age <= 60: age_groups_data['51-60'] += item['count']
            else: age_groups_data['>60'] += item['count']
        
        # توزیع گروه خونی
        blood_type_distribution = filtered_patients_qs.values('blood_type').annotate(count=Count('id')).order_by('blood_type')
        blood_type_chart_labels = [dict(BLOOD_TYPE_CHOICES).get(item['blood_type'], 'نامشخص') for item in blood_type_distribution]
        blood_type_chart_data = [item['count'] for item in blood_type_distribution]

        # توزیع شرکت‌ها
        company_distribution = filtered_patients_qs.values('company__name').annotate(count=Count('id')).order_by('company__name')
        company_chart_labels = [item['company__name'] if item['company__name'] else 'نامشخص' for item in company_distribution]
        company_chart_data = [item['count'] for item in company_distribution]

        # روند ثبت‌نام ماهانه
        registration_trend_monthly = filtered_patients_qs.annotate(month=TruncMonth('registered_at')).values('month').annotate(count=Count('id')).order_by('month')
        monthly_trend_labels = [JalaliDate(item['month']).strftime('%B %Y') for item in registration_trend_monthly if item['month']]
        monthly_trend_data = [item['count'] for item in registration_trend_monthly]

        # جدید: روند ثبت‌نام روزانه
        registration_trend_daily = filtered_patients_qs.annotate(day=TruncDay('registered_at')).values('day').annotate(count=Count('id')).order_by('day')
        daily_trend_labels = [JalaliDate(item['day']).strftime('%Y/%m/%d') for item in registration_trend_daily if item['day']]
        daily_trend_data = [item['count'] for item in registration_trend_daily]

        context.update({
            'patients': patients_with_visits,
            'total_patients_count': filtered_patients_qs.count(),
            'gender_chart_labels': json.dumps(gender_chart_labels, ensure_ascii=False),
            'gender_chart_data': json.dumps(gender_chart_data),
            'age_chart_labels': json.dumps(list(age_groups_data.keys()), ensure_ascii=False),
            'age_chart_data': json.dumps(list(age_groups_data.values())),
            'blood_type_chart_labels': json.dumps(blood_type_chart_labels, ensure_ascii=False),
            'blood_type_chart_data': json.dumps(blood_type_chart_data),
            'company_chart_labels': json.dumps(company_chart_labels, ensure_ascii=False),
            'company_chart_data': json.dumps(company_chart_data),
            'monthly_trend_labels': json.dumps(monthly_trend_labels, ensure_ascii=False),
            'monthly_trend_data': json.dumps(monthly_trend_data),
            'daily_trend_labels': json.dumps(daily_trend_labels, ensure_ascii=False),
            'daily_trend_data': json.dumps(daily_trend_data),
        })
    else:
        # اگر فیلتر اعمال نشده، مقادیر پیش‌فرض و خالی را به context اضافه کن
        context.update({
            'patients': [],
            'total_patients_count': 0,
            'gender_chart_labels': '[]', 'gender_chart_data': '[]',
            'age_chart_labels': '[]', 'age_chart_data': '[]',
            'blood_type_chart_labels': '[]', 'blood_type_chart_data': '[]',
            'company_chart_labels': '[]', 'company_chart_data': '[]',
            'monthly_trend_labels': '[]', 'monthly_trend_data': '[]',
            'daily_trend_labels': '[]', 'daily_trend_data': '[]',
        })

    return render(request, 'reports/patient_report.html', context)




@login_required
def generic_visit_report_view(request):
    report_title = "گزارش عمومی ویزیت‌ها"
    
    print(f"\n--- DEBUG: {report_title} View ---")
    print(f"Request GET parameters: {request.GET}")

    visit_filter = VisitFilter(request.GET, queryset=Visit.objects.select_related('patient', 'doctor', 'reason_for_visit', 'treatment_result', 'patient__company').all())
    filtered_visits_qs = visit_filter.qs

    print(f"Filtered Visits Query (SQL): {filtered_visits_qs.query}")
    print(f"Filtered Visits Count: {filtered_visits_qs.count()}")

    context = base_report_context(request, report_title, filter_form=visit_filter.form)

    is_filtered = bool(request.GET)
    context['is_filtered'] = is_filtered

    if is_filtered:
        total_visits_count = filtered_visits_qs.count()

        # نمودار و جدول تعداد مراجعه بر اساس شرکت‌ها
        company_visit_distribution = filtered_visits_qs.values('patient__company__name').annotate(
            count=Count('id')
        ).order_by('patient__company__name')
        company_visit_labels = [item['patient__company__name'] if item['patient__company__name'] else 'نامشخص' for item in company_visit_distribution]
        company_visit_data = [item['count'] for item in company_visit_distribution]
        company_table_data = list(company_visit_distribution) # برای جدول

        # نمودار و جدول مراجعه بر اساس علت مراجعه
        reason_distribution = filtered_visits_qs.values('reason_for_visit__name').annotate(
            count=Count('id')
        ).order_by('reason_for_visit__name')
        reason_labels = [item['reason_for_visit__name'] if item['reason_for_visit__name'] else 'نامشخص' for item in reason_distribution]
        reason_data = [item['count'] for item in reason_distribution]
        reason_table_data = list(reason_distribution) # برای جدول

        # نمودار و جدول نتیجه درمان
        treatment_distribution = filtered_visits_qs.values('treatment_result__name').annotate(
            count=Count('id')
        ).order_by('treatment_result__name')
        treatment_labels = [item['treatment_result__name'] if item['treatment_result__name'] else 'نامشخص' for item in treatment_distribution]
        treatment_data = [item['count'] for item in treatment_distribution]
        treatment_table_data = list(treatment_distribution) # برای جدول

        # نمودار و جدول نوع حادثه
        incident_type_distribution = filtered_visits_qs.values('incident_type').annotate(
            count=Count('id')
        ).order_by('incident_type')
        
        # تبدیل INCIDENT_TYPE_CHOICES به دیکشنری برای استفاده راحت‌تر
        dict_incident_type_choices = dict(INCIDENT_TYPE_CHOICES)

        incident_type_labels = [dict_incident_type_choices.get(item['incident_type'], 'نامشخص') for item in incident_type_distribution]
        incident_type_data = [item['count'] for item in incident_type_distribution]
        
        # **آماده‌سازی داده‌های جدول با نام قابل نمایش (display name) از پیش محاسبه شده**
        incident_type_table_data = []
        for item in incident_type_distribution:
            display_name = dict_incident_type_choices.get(item['incident_type'], 'نامشخص')
            incident_type_table_data.append({
                'incident_type_display': display_name, # نام قابل نمایش
                'incident_type_code': item['incident_type'], # کد اصلی (اختیاری)
                'count': item['count']
            })

        # نمودار روند ماهانه حادثه (فقط نمودار)
        monthly_incident_trend = filtered_visits_qs.exclude(incident_type='none').annotate(
            month=TruncMonth('visit_date')
        ).values('month', 'incident_type').annotate(
            count=Count('id')
        ).order_by('month', 'incident_type')

        incident_trend_data_processed = collections.defaultdict(lambda: collections.defaultdict(int))
        all_months = sorted(list(set([item['month'] for item in monthly_incident_trend if item['month']])))
        
        for item in monthly_incident_trend:
            if item['month']:
                incident_trend_data_processed[item['month']][item['incident_type']] = item['count']
        
        monthly_incident_labels = [JalaliDate.to_jalali(m).strftime('%B %Y') for m in all_months]
        
        incident_trend_datasets = []
        for incident_type_code, incident_type_display in INCIDENT_TYPE_CHOICES:
            if incident_type_code != 'none':
                data_for_type = [incident_trend_data_processed[month][incident_type_code] for month in all_months]
                # استفاده از هش برای رنگ‌های منحصر به فرد اما ثابت
                hash_val = hash(incident_type_code)
                color_r = (hash_val & 0xFF0000) >> 16
                color_g = (hash_val & 0x00FF00) >> 8
                color_b = (hash_val & 0x0000FF)
                incident_trend_datasets.append({
                    'label': incident_type_display,
                    'data': data_for_type,
                    'borderColor': f'rgba({color_r}, {color_g}, {color_b}, 1)',
                    'backgroundColor': f'rgba({color_r}, {color_g}, {color_b}, 0.2)',
                    'tension': 0.1,
                    'fill': False
                })

        # نمودار روند ثبت ویزیت روزانه و ماهانه (فقط نمودار)
        # روند ماهانه
        monthly_visit_trend = filtered_visits_qs.annotate(
            month=TruncMonth('visit_date')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')

        monthly_visit_labels = []
        monthly_visit_data = []
        for item in monthly_visit_trend:
            if item['month']:
                jalali_month = JalaliDate.to_jalali(item['month']).strftime('%B %Y')
                monthly_visit_labels.append(jalali_month)
                monthly_visit_data.append(item['count'])

        # روند روزانه
        daily_visit_trend = filtered_visits_qs.annotate(
            day=TruncDay('visit_date')
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day')

        daily_visit_labels = []
        daily_visit_data = []
        for item in daily_visit_trend:
            if item['day']:
                jalali_day = JalaliDate.to_jalali(item['day']).strftime('%Y/%m/%d')
                daily_visit_labels.append(jalali_day)
                daily_visit_data.append(item['count'])
        
        # نمودار هیستوگرام فشار خون ویزیت شده‌ها (فقط نمودار)
        blood_pressures = filtered_visits_qs.exclude(blood_pressure__isnull=True).values_list('blood_pressure', flat=True)
        systolic_pressures = []
        for bp_str in blood_pressures:
            try:
                parts = bp_str.split('/')
                if len(parts) == 2:
                    systolic_pressures.append(int(parts[0]))
            except ValueError:
                continue
        
        systolic_bins = collections.defaultdict(int)
        for sp in systolic_pressures:
            if sp < 90: systolic_bins['< 90'] += 1
            elif 90 <= sp <= 120: systolic_bins['90-120'] += 1
            elif 121 <= sp <= 140: systolic_bins['121-140'] += 1
            elif 141 <= sp <= 160: systolic_bins['141-160'] += 1
            else: systolic_bins['> 160'] += 1
        
        # مرتب‌سازی کلیدها به صورت سفارشی
        systolic_labels_order = ['< 90', '90-120', '121-140', '141-160', '> 160']
        systolic_labels = [label for label in systolic_labels_order if label in systolic_bins]
        systolic_data = [systolic_bins[label] for label in systolic_labels]

        # نمودار سنی ویزیت شده‌ها (فقط نمودار)
        age_at_visit_data = []
        for visit in filtered_visits_qs.filter(patient__date_of_birth__isnull=False):
            if visit.patient.date_of_birth:
                age_td = visit.visit_date.date() - visit.patient.date_of_birth
                age_years = age_td.days // 365
                age_at_visit_data.append(age_years)
        
        age_bins = {'0-10': 0, '11-20': 0, '21-30': 0, '31-40': 0, '41-50': 0, '51-60': 0, '>60': 0}
        for age in age_at_visit_data:
            if 0 <= age <= 10: age_bins['0-10'] += 1
            elif 11 <= age <= 20: age_bins['11-20'] += 1
            elif 21 <= age <= 30: age_bins['21-30'] += 1
            elif 31 <= age <= 40: age_bins['31-40'] += 1
            elif 41 <= age <= 50: age_bins['41-50'] += 1
            elif 51 <= age <= 60: age_bins['51-60'] += 1
            else: age_bins['>60'] += 1

        age_at_visit_labels = list(age_bins.keys())
        age_at_visit_data = list(age_bins.values())

        # نمودار BMI (فقط نمودار)
        bmi_data_for_chart = []
        for visit in filtered_visits_qs.filter(height_cm__isnull=False, weight_kg__isnull=False):
            if visit.bmi is not None:
                bmi_data_for_chart.append(float(visit.bmi))

        bmi_bins = collections.defaultdict(int)
        for bmi_val in bmi_data_for_chart:
            if bmi_val < 18.5: bmi_bins['کم وزن (< 18.5)'] += 1
            elif 18.5 <= bmi_val <= 24.9: bmi_bins['وزن طبیعی (18.5-24.9)'] += 1
            elif 25.0 <= bmi_val <= 29.9: bmi_bins['اضافه وزن (25.0-29.9)'] += 1
            elif 30.0 <= bmi_val <= 34.9: bmi_bins['چاقی درجه ۱ (30.0-34.9)'] += 1
            elif 35.0 <= bmi_val <= 39.9: bmi_bins['چاقی درجه ۲ (35.0-39.9)'] += 1
            else: bmi_bins['چاقی درجه ۳ (>= 40.0)'] += 1

        bmi_labels_ordered = [
            'کم وزن (< 18.5)', 'وزن طبیعی (18.5-24.9)', 'اضافه وزن (25.0-29.9)',
            'چاقی درجه ۱ (30.0-34.9)', 'چاقی درجه ۲ (35.0-39.9)', 'چاقی درجه ۳ (>= 40.0)'
        ]
        bmi_chart_labels = [label for label in bmi_labels_ordered if label in bmi_bins]
        bmi_chart_data = [bmi_bins[label] for label in bmi_chart_labels]


        context.update({
            'total_visits_count': total_visits_count,

            # Chart Data
            'company_visit_chart_labels': json.dumps(company_visit_labels, ensure_ascii=False),
            'company_visit_chart_data': json.dumps(company_visit_data),
            'reason_chart_labels': json.dumps(reason_labels, ensure_ascii=False),
            'reason_chart_data': json.dumps(reason_data),
            'treatment_chart_labels': json.dumps(treatment_labels, ensure_ascii=False),
            'treatment_chart_data': json.dumps(treatment_data),
            'incident_type_chart_labels': json.dumps(incident_type_labels, ensure_ascii=False),
            'incident_type_chart_data': json.dumps(incident_type_data),
            'monthly_incident_labels': json.dumps(monthly_incident_labels, ensure_ascii=False),
            'incident_trend_datasets': json.dumps(incident_trend_datasets, ensure_ascii=False),
            'monthly_visit_labels': json.dumps(monthly_visit_labels, ensure_ascii=False),
            'monthly_visit_data': json.dumps(monthly_visit_data),
            'daily_visit_labels': json.dumps(daily_visit_labels, ensure_ascii=False),
            'daily_visit_data': json.dumps(daily_visit_data),
            'systolic_labels': json.dumps(systolic_labels, ensure_ascii=False),
            'systolic_data': json.dumps(systolic_data),
            'age_at_visit_labels': json.dumps(age_at_visit_labels, ensure_ascii=False),
            'age_at_visit_data': json.dumps(age_at_visit_data),
            'bmi_chart_labels': json.dumps(bmi_chart_labels, ensure_ascii=False),
            'bmi_chart_data': json.dumps(bmi_chart_data),

            # Table Data (updated to include display names where applicable)
            'company_table_data': company_table_data,
            'reason_table_data': reason_table_data,
            'treatment_table_data': treatment_table_data,
            'incident_type_table_data': incident_type_table_data, # این لیست اکنون حاوی 'incident_type_display' است.

            'visits': filtered_visits_qs,
        })
    else:
        context.update({
            'total_visits_count': 0,
            # Empty Chart Data
            'company_visit_chart_labels': '[]', 'company_visit_chart_data': '[]',
            'reason_chart_labels': '[]', 'reason_chart_data': '[]',
            'treatment_chart_labels': '[]', 'treatment_chart_data': '[]',
            'incident_type_chart_labels': '[]', 'incident_type_chart_data': '[]',
            'monthly_incident_labels': '[]', 'incident_trend_datasets': '[]',
            'monthly_visit_labels': '[]', 'monthly_visit_data': '[]',
            'daily_visit_labels': '[]', 'daily_visit_data': '[]',
            'systolic_labels': '[]', 'systolic_data': '[]',
            'age_at_visit_labels': '[]', 'age_at_visit_data': '[]',
            'bmi_chart_labels': '[]', 'bmi_chart_data': '[]',
            # Empty Table Data
            'company_table_data': [],
            'reason_table_data': [],
            'treatment_table_data': [],
            'incident_type_table_data': [],
            'visits': Visit.objects.none(),
        })

    return render(request, 'reports/generic_visit_report.html', context)

@login_required
def drug_report_view(request):
    report_title = "گزارش داروها"
    
    is_filtered = bool(request.GET)

    # 1. ابتدا DrugFilter را روی کل داروها اعمال می‌کنیم.
    drug_filter = DrugFilter(request.GET, queryset=Drug.objects.all())
    base_filtered_drugs_qs = drug_filter.qs # این کوئری فیلترهای DrugFilter را اعمال کرده است.

    # 2. حالا از یک Subquery برای محاسبه موجودی کل هر دارو استفاده می‌کنیم.
    # این Subquery مجموع quantity بچ‌های مربوط به هر دارو را به دست می‌آورد.
    batch_sum_subquery = DrugBatch.objects.filter(
        drug_id=OuterRef('pk') # این خط Subquery را به کوئری اصلی Drug مرتبط می‌کند.
    ).values('drug').annotate(
        total_quantity=Coalesce(Sum('quantity'), 0) # مجموع موجودی، اگر بچی نبود 0 را برگردان
    ).values('total_quantity') # فقط مقدار مجموع را انتخاب کن.

    # 3. سپس stock_quantity را با استفاده از Subquery به کوئری اصلی اضافه می‌کنیم.
    # Subquery تضمین می‌کند که Sum به درستی و بدون تکرار (ناشی از JOIN) محاسبه شود.
    filtered_drugs_with_stock = base_filtered_drugs_qs.annotate(
        stock_quantity=Subquery(batch_sum_subquery, output_field=fields.IntegerField())
    ).order_by('name')

    # اگر Drug با هیچ بچی مرتبط نباشد، Subquery ممکن است None برگرداند.
    # Coalesce در Subquery (بالا) این مورد را مدیریت می‌کند، اما این خط برای اطمینان بیشتر است.
    filtered_drugs_with_stock = filtered_drugs_with_stock.annotate(
        stock_quantity=Coalesce('stock_quantity', 0)
    )

    context = base_report_context(request, report_title, filter_form=drug_filter.form)
    context['is_filtered'] = is_filtered

    if is_filtered:
        show_batch_details = request.GET.get('show_batch_details') == 'on'
        
        # برای بهینه‌سازی، اگر جزئیات بچ نیاز است، آنها را prefetch می‌کنیم
        if show_batch_details:
            # از filtered_drugs_with_stock که قبلاً annotate و فیلتر شده، استفاده می‌کنیم
            drug_inventory_data = filtered_drugs_with_stock.prefetch_related('batches') # 'batches__supplier' اگر نیاز به supplier بچ هست
        else:
            # اگر جزئیات بچ نمایش داده نمی‌شود، همین کوئری کافی است
            drug_inventory_data = filtered_drugs_with_stock 

        # شناسه‌های داروهای فیلتر شده برای استفاده در کوئری‌های بعدی (خرید و مصرف)
        # از filtered_drugs_with_stock استفاده می‌کنیم که صحیحاً فیلتر شده است
        filtered_drug_ids = filtered_drugs_with_stock.values_list('id', flat=True)

        # آمار داروهای خریداری شده
        purchase_items = PurchaseInvoiceItem.objects.filter(drug__id__in=filtered_drug_ids)
        total_purchased_quantity_overall = purchase_items.aggregate(total=Sum('quantity'))['total'] or 0
        top_5_purchased_drugs = purchase_items.values('drug__name').annotate(
            total_purchased_quantity=Sum('quantity')
        ).order_by('-total_purchased_quantity')[:5]
        least_purchased_drug = purchase_items.values('drug__name').annotate(
            total_purchased_quantity=Sum('quantity')
        ).order_by('total_purchased_quantity').first()

        # آمار داروهای مصرف شده
        consumed_items = VisitItem.objects.filter(drug__id__in=filtered_drug_ids)
        total_consumed_quantity_overall = consumed_items.aggregate(total=Sum('quantity'))['total'] or 0
        top_5_consumed_drugs = consumed_items.values('drug__name').annotate(
            total_consumed_quantity=Sum('quantity')
        ).order_by('-total_consumed_quantity')[:5]
        least_consumed_drug = consumed_items.values('drug__name').annotate(
            total_consumed_quantity=Sum('quantity')
        ).order_by('total_consumed_quantity').first()

        context.update({
            'drug_inventory_data': drug_inventory_data,
            'total_drugs_count': filtered_drugs_with_stock.count(), # شمارش داروها بعد از محاسبه صحیح موجودی
            'show_batch_details': show_batch_details,
            
            # آمار خرید
            'total_purchased_quantity_overall': total_purchased_quantity_overall,
            'top_5_purchased_drugs': top_5_purchased_drugs,
            'least_purchased_drug': least_purchased_drug,
            
            # آمار مصرف
            'total_consumed_quantity_overall': total_consumed_quantity_overall,
            'top_5_consumed_drugs': top_5_consumed_drugs,
            'least_consumed_drug': least_consumed_drug,
        })
    else:
        # مقداردهی اولیه context در صورتی که فیلتری اعمال نشده باشد
        context.update({
            'drug_inventory_data': [],
            'total_drugs_count': 0,
            'show_batch_details': False,
            'total_purchased_quantity_overall': 0,
            'top_5_purchased_drugs': [],
            'least_purchased_drug': None,
            'total_consumed_quantity_overall': 0,
            'top_5_consumed_drugs': [],
            'least_consumed_drug': None,
        })

    return render(request, 'reports/drug_report.html', context)

@login_required
def company_visit_report_view(request):
    report_title = "گزارش ویزیت شرکت‌ها"
    from visits.views import company_visit_report_view as visits_company_visit_report_view
    return visits_company_visit_report_view(request, template_name='reports/company_visit_report.html')


@login_required
def get_report_data(report_type, company_id, final_opinion_filters, re_exam_date_jalali_str):
    patients_without_exams = None
    patients_without_personnel_id = None
    patients_due_for_re_exam = None
    final_opinion_exams = None
    personnel_with_code = None
    exams_without_personnel_id = None
    personnel_without_exams = None
    
    report_title = ""
    re_exam_target_date_jalali = None
    
    base_query = Patient.objects.all()
    if company_id:
        base_query = base_query.filter(company_id=company_id)
        
    patients_with_last_exam = {
        item['patient_id']: item['last_exam_date']
        for item in PeriodicExamination.objects.values('patient_id')
                                                 .annotate(last_exam_date=Max('exam_date'))
    }

    if report_type == 'no_exams':
        patients_without_exams = base_query.filter(periodic_examinations__isnull=True).order_by('first_name', 'last_name')
        for patient in patients_without_exams:
            last_exam_date = patients_with_last_exam.get(patient.id)
            if last_exam_date:
                patient.last_exam_date_jalali = jdatetime.date.fromgregorian(date=last_exam_date)
            else:
                patient.last_exam_date_jalali = None
        report_title = "بیمارانی که معاینه دوره‌ای ندارند"
    
    elif report_type == 'no_personnel_id':
        patients_without_personnel_id = base_query.filter(
            Q(personnel_number__isnull=True) | Q(personnel_number__exact='')
        ).order_by('first_name', 'last_name')
        for patient in patients_without_personnel_id:
            last_exam_date = patients_with_last_exam.get(patient.id)
            if last_exam_date:
                patient.last_exam_date_jalali = jdatetime.date.fromgregorian(date=last_exam_date)
            else:
                patient.last_exam_date_jalali = None
        report_title = "بیمارانی که کد پرسنلی ندارند"

    elif report_type == 'due_for_re_exam':
        if re_exam_date_jalali_str:
            try:
                j_date = jdatetime.datetime.strptime(re_exam_date_jalali_str, '%Y/%m/%d').date()
                re_exam_target_date_gregorian = j_date.togregorian()
                re_exam_target_date_jalali = j_date
            except (ValueError, ImportError):
                re_exam_target_date_gregorian = None
                re_exam_target_date_jalali = None

            if re_exam_target_date_gregorian:
                latest_exams = PeriodicExamination.objects.filter(patient__in=base_query).values('patient').annotate(
                    last_exam_date=Max('exam_date')
                ).order_by('patient')
                
                patients_due_for_re_exam = []
                for le in latest_exams:
                    patient_id = le['patient']
                    last_exam_date = le['last_exam_date']

                    if last_exam_date:
                        expiry_date_gregorian = last_exam_date + timedelta(days=365)
                        if expiry_date_gregorian <= re_exam_target_date_gregorian:
                            patient_obj = Patient.objects.get(id=patient_id)
                            last_exam_date_jalali = jdatetime.date.fromgregorian(date=last_exam_date)
                            expiry_date_jalali = jdatetime.date.fromgregorian(date=expiry_date_gregorian)
                            patients_due_for_re_exam.append({
                                'patient': patient_obj,
                                'last_exam_date_jalali': last_exam_date_jalali,
                                'expiry_date_jalali': expiry_date_jalali
                            })
                report_title = f"بیمارانی که نیاز به معاینه مجدد دارند (تا تاریخ: {re_exam_target_date_jalali})"

    elif report_type == 'final_opinion':
        opinion_query = Q()
        if 'conditional' in final_opinion_filters:
            opinion_query |= Q(final_opinion_text__icontains='مشروط')
        if 'unconditional' in final_opinion_filters:
            opinion_query |= Q(final_opinion_text__icontains='بلامانع')
        if 'not_declared' in final_opinion_filters:
            opinion_query |= Q(final_opinion_text__isnull=True) | Q(final_opinion_text__exact='') | Q(final_opinion_text__in=['-', '.'])
        # اضافه کردن فیلتر جدید
        if 'waiting_for_result' in final_opinion_filters:
            opinion_query |= Q(final_opinion_text__icontains='در انتظار نتیجه')
            
        filtered_exams = PeriodicExamination.objects.filter(opinion_query, patient__in=base_query).select_related('patient').order_by('-exam_date')

        for exam in filtered_exams:
            if exam.exam_date:
                exam.exam_date_jalali = jdatetime.date.fromgregorian(date=exam.exam_date)
            else:
                exam.exam_date_jalali = None
        final_opinion_exams = filtered_exams
        
        filter_names = {
            'conditional': 'مشروط', 
            'unconditional': 'بلامانع', 
            'not_declared': 'عدم اعلام نظر',
            'waiting_for_result': 'در انتظار نتیجه',
        }
        selected_names = [filter_names[f] for f in final_opinion_filters if f in filter_names]
        report_title = f"معاینات با نظریه نهایی: {', '.join(selected_names)}" if selected_names else "معاینات با نظریه نهایی"

    # گزارش جدید: بیمارانی که کد پرسنلی دارند
    elif report_type == 'personnel_with_code':
        personnel_with_code = base_query.filter(
            Q(personnel_number__isnull=False) & ~Q(personnel_number__exact='')
        ).order_by('personnel_number')
        
        report_title = "بیمارانی که کد پرسنلی دارند (با نظریه نهایی)"
        
        for patient in personnel_with_code:
            try:
                last_exam = PeriodicExamination.objects.filter(patient=patient).latest('exam_date')
                
                # اضافه کردن تاریخ معاینه
                patient.last_exam_date_jalali = jdatetime.date.fromgregorian(date=last_exam.exam_date)

                # بررسی و جایگزینی نظریه نهایی
                if last_exam.final_opinion_text == '-' or last_exam.final_opinion_text == '.' or last_exam.final_opinion_text is None:
                    patient.final_opinion_text = "منتظر اعلام نتیجه"
                    patient.final_opinion_conditions = "بیمار جهت بررسی بیشتر به متخصص ارجاع شده است اما تاکنون نتیجه را ارائه نکرده است"
                else:
                    patient.final_opinion_text = last_exam.final_opinion_text
                    patient.final_opinion_conditions = last_exam.final_opinion_conditions

            except PeriodicExamination.DoesNotExist:
                # اگر معاینه‌ای وجود نداشت
                patient.last_exam_date_jalali = "-"
                patient.final_opinion_text = "طب کار ندارد"
                patient.final_opinion_conditions = "-"

    # گزارش جدید: بیمارانی که معاینه دوره‌ای دارند ولی کد پرسنلی ندارند
    elif report_type == 'exams_without_personnel_id':
        exams_without_personnel_id = base_query.filter(
            (Q(personnel_number__isnull=True) | Q(personnel_number__exact='')),
            periodic_examinations__isnull=False
        ).distinct().order_by('first_name', 'last_name')
        report_title = "بیمارانی که معاینه دارند ولی کد پرسنلی ندارند"
        for patient in exams_without_personnel_id:
            last_exam_date = patients_with_last_exam.get(patient.id)
            if last_exam_date:
                patient.last_exam_date_jalali = jdatetime.date.fromgregorian(date=last_exam_date)
            else:
                patient.last_exam_date_jalali = None
                
    # گزارش جدید: بیمارانی که کد پرسنلی دارند ولی معاینه دوره‌ای ندارند
    elif report_type == 'personnel_without_exams':
        personnel_without_exams = base_query.filter(
            Q(personnel_number__isnull=False) & ~Q(personnel_number__exact=''),
            periodic_examinations__isnull=True
        ).order_by('personnel_number')
        report_title = "بیمارانی که کد پرسنلی دارند معاینه ندارند"
        for patient in personnel_without_exams:
            last_exam_date = patients_with_last_exam.get(patient.id)
            if last_exam_date:
                patient.last_exam_date_jalali = jdatetime.date.fromgregorian(date=last_exam_date)
            else:
                patient.last_exam_date_jalali = None

    return {
        'patients_without_exams': patients_without_exams,
        'patients_without_personnel_id': patients_without_personnel_id,
        'patients_due_for_re_exam': patients_due_for_re_exam,
        'final_opinion_exams': final_opinion_exams,
        'personnel_with_code': personnel_with_code,
        'exams_without_personnel_id': exams_without_personnel_id,
        'personnel_without_exams': personnel_without_exams,
        're_exam_target_date_jalali': re_exam_target_date_jalali,
        'report_title': report_title,
    }

@login_required
def reports_view(request):
    report_type = request.GET.get('report_type')
    company_id = request.GET.get('company')
    final_opinion_filters = request.GET.getlist('final_opinion_filter')
    re_exam_date_jalali_str = request.GET.get('re_exam_date_jalali')
    
    context = get_report_data(report_type, company_id, final_opinion_filters, re_exam_date_jalali_str)
    context['selected_report_type'] = report_type
    context['companies'] = Company.objects.all()
    context['selected_company'] = company_id
    context['selected_final_opinion_filters'] = final_opinion_filters
    
    return render(request, 'lab_results/reports.html', context)
@login_required
def export_excel(request):
    report_type = request.GET.get('report_type')
    company_id = request.GET.get('company')
    final_opinion_filters = request.GET.getlist('final_opinion_filter')
    re_exam_date_jalali_str = request.GET.get('re_exam_date_jalali')
    
    report_data = get_report_data(report_type, company_id, final_opinion_filters, re_exam_date_jalali_str)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_data["report_title"]}.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "گزارش"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    def format_header(row):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if report_type == 'personnel_with_code':
        headers = ["ردیف", "کد پرسنلی", "نام کامل", "کد ملی", "شغل درخواستی", "تاریخ آزمایش", "نظریه نهایی", "شروط نظریه نهایی"]
        worksheet.append(headers)
        format_header(worksheet[1])
        
        data_list = report_data['personnel_with_code']
        for index, patient in enumerate(data_list, 1):
            last_exam_date_str = patient.last_exam_date_jalali
            if isinstance(patient.last_exam_date_jalali, jdatetime.date):
                last_exam_date_str = patient.last_exam_date_jalali.strftime('%Y/%m/%d')
            
            worksheet.append([
                index,
                patient.personnel_number or "-",
                f"{patient.first_name} {patient.last_name}",
                patient.national_code or "-",
                patient.occupation or "-",
                last_exam_date_str,
                patient.final_opinion_text or "-",
                patient.final_opinion_conditions or "-",
            ])
            
    # افزودن منطق برای سایر انواع گزارشات
    elif report_type == 'no_exams':
        headers = ["ردیف", "نام کامل", "کد ملی", "کد پرسنلی", "شماره تماس", "شغل درخواستی", "آخرین تاریخ معاینه"]
        worksheet.append(headers)
        format_header(worksheet[1])
        
        data_list = report_data['patients_without_exams']
        for index, patient in enumerate(data_list, 1):
            last_exam_date_str = patient.last_exam_date_jalali
            if isinstance(patient.last_exam_date_jalali, jdatetime.date):
                last_exam_date_str = patient.last_exam_date_jalali.strftime('%Y/%m/%d')
                
            worksheet.append([
                index,
                f"{patient.first_name} {patient.last_name}",
                patient.national_code or "-",
                patient.personnel_number or "-",
                patient.phone_number or "-",
                patient.occupation or "-",
                last_exam_date_str,
            ])

    elif report_type == 'final_opinion':
        headers = ["ردیف", "نام بیمار", "کد پرسنلی", "تاریخ معاینه", "نظریه نهایی", "شروط نظریه نهایی", "شغل درخواستی"]
        worksheet.append(headers)
        format_header(worksheet[1])
        
        data_list = report_data['final_opinion_exams']
        for index, exam in enumerate(data_list, 1):
            exam_date_str = exam.exam_date_jalali
            if isinstance(exam.exam_date_jalali, jdatetime.date):
                exam_date_str = exam.exam_date_jalali.strftime('%Y/%m/%d')
            
            worksheet.append([
                index,
                f"{exam.patient.first_name} {exam.patient.last_name}",
                exam.patient.personnel_number or "-",
                exam_date_str,
                exam.final_opinion_text or "-",
                exam.final_opinion_conditions or "-",
                exam.patient.occupation or "-",
            ])

    elif report_type == 'no_personnel_id':
        headers = ["ردیف", "نام کامل", "کد ملی", "کد پرسنلی", "شماره تماس", "شغل درخواستی", "آخرین تاریخ معاینه"]
        worksheet.append(headers)
        format_header(worksheet[1])
        
        data_list = report_data['patients_without_personnel_id']
        for index, patient in enumerate(data_list, 1):
            last_exam_date_str = patient.last_exam_date_jalali
            if isinstance(patient.last_exam_date_jalali, jdatetime.date):
                last_exam_date_str = patient.last_exam_date_jalali.strftime('%Y/%m/%d')
            
            worksheet.append([
                index,
                f"{patient.first_name} {patient.last_name}",
                patient.national_code or "-",
                patient.personnel_number or "-",
                patient.phone_number or "-",
                patient.occupation or "-",
                last_exam_date_str,
            ])

    elif report_type == 'exams_without_personnel_id':
        headers = ["ردیف", "نام کامل", "کد ملی", "شغل درخواستی", "شماره تماس", "آخرین تاریخ معاینه"]
        worksheet.append(headers)
        format_header(worksheet[1])
        
        data_list = report_data['exams_without_personnel_id']
        for index, patient in enumerate(data_list, 1):
            last_exam_date_str = patient.last_exam_date_jalali
            if isinstance(patient.last_exam_date_jalali, jdatetime.date):
                last_exam_date_str = patient.last_exam_date_jalali.strftime('%Y/%m/%d')
            
            worksheet.append([
                index,
                f"{patient.first_name} {patient.last_name}",
                patient.national_code or "-",
                patient.occupation or "-",
                patient.phone_number or "-",
                last_exam_date_str,
            ])

    elif report_type == 'personnel_without_exams':
        headers = ["ردیف", "کد پرسنلی", "نام کامل", "کد ملی", "شغل درخواستی", "شماره تماس"]
        worksheet.append(headers)
        format_header(worksheet[1])
        
        data_list = report_data['personnel_without_exams']
        for index, patient in enumerate(data_list, 1):
            worksheet.append([
                index,
                patient.personnel_number or "-",
                f"{patient.first_name} {patient.last_name}",
                patient.national_code or "-",
                patient.occupation or "-",
                patient.phone_number or "-",
            ])

    elif report_type == 'due_for_re_exam':
        headers = ["ردیف", "نام کامل", "کد پرسنلی", "شغل درخواستی", "آخرین تاریخ معاینه", "تاریخ اعتبار آزمایش", "شماره تماس"]
        worksheet.append(headers)
        format_header(worksheet[1])
        
        data_list = report_data['patients_due_for_re_exam']
        for index, item in enumerate(data_list, 1):
            last_exam_date_str = item['last_exam_date_jalali']
            if isinstance(item['last_exam_date_jalali'], jdatetime.date):
                last_exam_date_str = item['last_exam_date_jalali'].strftime('%Y/%m/%d')
            
            expiry_date_str = item['expiry_date_jalali']
            if isinstance(item['expiry_date_jalali'], jdatetime.date):
                expiry_date_str = item['expiry_date_jalali'].strftime('%Y/%m/%d')
            
            worksheet.append([
                index,
                f"{item['patient'].first_name} {item['patient'].last_name}",
                item['patient'].personnel_number or "-",
                item['patient'].occupation or "-",
                last_exam_date_str,
                expiry_date_str,
                item['patient'].phone_number or "-",
            ])

    workbook.save(response)
    return response

@login_required
def reports_view(request):
    report_type = request.GET.get('report_type')
    company_id = request.GET.get('company')
    final_opinion_filters = request.GET.getlist('final_opinion_filter')
    re_exam_date_jalali_str = request.GET.get('re_exam_date_jalali')
    
    context = get_report_data(report_type, company_id, final_opinion_filters, re_exam_date_jalali_str)
    context['selected_report_type'] = report_type
    context['companies'] = Company.objects.all()
    context['selected_company'] = company_id
    context['selected_final_opinion_filters'] = final_opinion_filters
    
    return render(request, 'lab_results/reports.html', context)




import datetime as dt_module
from django.shortcuts import render
from django.db.models import Value, CharField, F, fields
from django.utils import timezone
from persiantools.jdatetime import JalaliDate, JalaliDateTime

# ۱. ایمپورت مدل‌ها در بالاترین سطح (خارج از تابع) برای جلوگیری از UnboundLocalError
# آدرس‌ها را بر اساس ساختار دقیق پروژه‌ات اصلاح کن
from drugs.models import Drug,  PurchaseInvoiceItem 

from visits.models import VisitItem
from .filters import DrugTransactionFilter
def get_drug_transactions_data(request):
    """تابع مشترک برای استخراج تراکنش‌ها در گزارش و چاپ"""
    drug_ids = request.GET.getlist('drug')
    select_all = request.GET.get('select_all_drugs') == 'on'
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    t_type_filter = request.GET.get('transaction_type', 'all')

    transactions = []
    selected_drugs = []
    start_dt = None
    end_dt = None

    # تبدیل تاریخ شمسی به میلادی
    try:
        table = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
        if start_date_str:
            p = [int(x) for x in start_date_str.translate(table).split('/')]
            start_dt = timezone.make_aware(dt_module.datetime.combine(JalaliDate(p[0],p[1],p[2]).to_gregorian(), dt_module.time.min))
        if end_date_str:
            p = [int(x) for x in end_date_str.translate(table).split('/')]
            end_dt = timezone.make_aware(dt_module.datetime.combine(JalaliDate(p[0],p[1],p[2]).to_gregorian(), dt_module.time.max))
    except: pass

    if select_all or drug_ids:
        if select_all:
            selected_drugs = Drug.objects.all()
            drug_ids = [int(d.id) for d in selected_drugs]
        else:
            drug_ids = [int(d_id) for d_id in drug_ids]
            selected_drugs = Drug.objects.filter(id__in=drug_ids)

        drug_balances = {}
        target_date = start_dt if start_dt else timezone.make_aware(dt_module.datetime(1900, 1, 1))
        
        for d_id in drug_ids:
            p_sum = PurchaseInvoiceItem.objects.filter(drug_id=d_id, invoice__invoice_date__lt=target_date).aggregate(s=models.Sum('quantity'))['s'] or 0
            v_sum = VisitItem.objects.filter(drug_id=d_id, visit__visit_date__lt=target_date).aggregate(s=models.Sum('quantity'))['s'] or 0
            drug_balances[d_id] = p_sum - v_sum

        # فیلتر خریدها
        p_filters = {'drug_id__in': drug_ids}
        if start_dt: p_filters['invoice__invoice_date__gte'] = start_dt
        if end_dt: p_filters['invoice__invoice_date__lte'] = end_dt

        purchases = []
        if t_type_filter in ['all', 'in']:
            purchases = list(PurchaseInvoiceItem.objects.filter(**p_filters).annotate(
                raw_date=F('invoice__invoice_date'),
                drug_name=F('drug__name'),
                q_in=F('quantity'),
                q_out=Value(0, output_field=fields.IntegerField()),
                op_type=Value('ورود', output_field=CharField())
            ).values('raw_date', 'drug_name', 'drug_id', 'q_in', 'q_out', 'op_type'))

        # فیلتر مصرف‌ها
        c_filters = {'drug_id__in': drug_ids}
        if start_dt: c_filters['visit__visit_date__gte'] = start_dt
        if end_dt: c_filters['visit__visit_date__lte'] = end_dt

        consumptions = []
        if t_type_filter in ['all', 'out']:
            consumptions = list(VisitItem.objects.filter(**c_filters).annotate(
                raw_date=F('visit__visit_date'),
                drug_name=F('drug__name'),
                q_in=Value(0, output_field=fields.IntegerField()),
                q_out=F('quantity'),
                op_type=Value('خروج', output_field=CharField())
            ).values('raw_date', 'drug_name', 'drug_id', 'q_in', 'q_out', 'op_type'))

        all_items = purchases + consumptions

        def safe_dt(val):
            if not val: return timezone.make_aware(dt_module.datetime(1900,1,1))
            if isinstance(val, dt_module.date) and not isinstance(val, dt_module.datetime):
                return timezone.make_aware(dt_module.datetime.combine(val, dt_module.time.min))
            return timezone.make_aware(val) if timezone.is_naive(val) else val

        all_items.sort(key=lambda x: safe_dt(x['raw_date']))

        for item in all_items:
            d_id = int(item['drug_id'])
            drug_balances[d_id] += (item['q_in'] - item['q_out'])
            item['balance_after'] = drug_balances[d_id]
            item['date_jalali'] = to_jalali_helper(item['raw_date'])

        transactions = all_items[::-1]
    
    return transactions, selected_drugs, bool(select_all or drug_ids)


def to_jalali_helper(dt_obj):
    if not dt_obj: return "-"
    try:
        if hasattr(dt_obj, 'hour'):
            dt_obj = timezone.localtime(dt_obj)
            return JalaliDateTime.to_jalali(dt_obj).strftime('%Y/%m/%d - %H:%M')
        return JalaliDate.to_jalali(dt_obj).strftime('%Y/%m/%d')
    except: return str(dt_obj)

def drug_transaction_report_view(request):
    drug_ids = request.GET.getlist('drug')
    select_all = request.GET.get('select_all_drugs') == 'on'
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    t_type_filter = request.GET.get('transaction_type', 'all')

    transactions = []
    selected_drugs = []
    start_dt = None
    end_dt = None

    # تبدیل تاریخ شمسی به میلادی
    try:
        table = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
        if start_date_str:
            p = [int(x) for x in start_date_str.translate(table).split('/')]
            start_dt = timezone.make_aware(dt_module.datetime.combine(JalaliDate(p[0],p[1],p[2]).to_gregorian(), dt_module.time.min))
        if end_date_str:
            p = [int(x) for x in end_date_str.translate(table).split('/')]
            end_dt = timezone.make_aware(dt_module.datetime.combine(JalaliDate(p[0],p[1],p[2]).to_gregorian(), dt_module.time.max))
    except: pass

    if select_all or drug_ids:
        if select_all:
            selected_drugs = Drug.objects.all()
            drug_ids = [int(d.id) for d in selected_drugs]
        else:
            # تبدیل آی‌دی‌ها به عدد برای جلوگیری از KeyError
            drug_ids = [int(d_id) for d_id in drug_ids]
            selected_drugs = Drug.objects.filter(id__in=drug_ids)

        # محاسبه موجودی اولیه هر دارو تا قبل از شروع بازه
        drug_balances = {}
        target_date = start_dt if start_dt else timezone.make_aware(dt_module.datetime(1900, 1, 1))
        
        for d_id in drug_ids:
            p_sum = PurchaseInvoiceItem.objects.filter(drug_id=d_id, invoice__invoice_date__lt=target_date).aggregate(s=models.Sum('quantity'))['s'] or 0
            v_sum = VisitItem.objects.filter(drug_id=d_id, visit__visit_date__lt=target_date).aggregate(s=models.Sum('quantity'))['s'] or 0
            drug_balances[d_id] = p_sum - v_sum

        # دریافت تراکنش‌ها
        p_filters = {'drug_id__in': drug_ids}
        if start_dt: p_filters['invoice__invoice_date__gte'] = start_dt
        if end_dt: p_filters['invoice__invoice_date__lte'] = end_dt

        purchases = []
        if t_type_filter in ['all', 'in']:
            purchases = list(PurchaseInvoiceItem.objects.filter(**p_filters).annotate(
                raw_date=F('invoice__invoice_date'),
                drug_name=F('drug__name'),
                q_in=F('quantity'),
                q_out=Value(0, output_field=fields.IntegerField()),
                op_type=Value('ورود', output_field=CharField())
            ).values('raw_date', 'drug_name', 'drug_id', 'q_in', 'q_out', 'op_type'))

        c_filters = {'drug_id__in': drug_ids}
        if start_dt: c_filters['visit__visit_date__gte'] = start_dt
        if end_dt: c_filters['visit__visit_date__lte'] = end_dt

        consumptions = []
        if t_type_filter in ['all', 'out']:
            consumptions = list(VisitItem.objects.filter(**c_filters).annotate(
                raw_date=F('visit__visit_date'),
                drug_name=F('drug__name'),
                q_in=Value(0, output_field=fields.IntegerField()),
                q_out=F('quantity'),
                op_type=Value('خروج', output_field=CharField())
            ).values('raw_date', 'drug_name', 'drug_id', 'q_in', 'q_out', 'op_type'))

        all_items = purchases + consumptions

        # تابع کمکی برای مرتب‌سازی (رفع باگ datetime vs date)
        def safe_dt(val):
            if not val: return timezone.make_aware(dt_module.datetime(1900,1,1))
            if isinstance(val, dt_module.date) and not isinstance(val, dt_module.datetime):
                return timezone.make_aware(dt_module.datetime.combine(val, dt_module.time.min))
            return timezone.make_aware(val) if timezone.is_naive(val) else val

        all_items.sort(key=lambda x: safe_dt(x['raw_date']))

        # محاسبه موجودی لحظه‌ای تفکیک شده
        for item in all_items:
            d_id = int(item['drug_id'])
            drug_balances[d_id] += (item['q_in'] - item['q_out'])
            item['balance_after'] = drug_balances[d_id]
            item['date_jalali'] = to_jalali_helper(item['raw_date'])

        transactions = all_items[::-1]
    today_jalali = JalaliDate.today().strftime('%Y/%m/%d')
    return render(request, 'reports/drug_transaction_report.html', {
        'report_title': "گزارش کارتکس و موجودی داروخانه",
        'drugs_list': Drug.objects.all(),
        'transactions': transactions,
        'selected_drugs': selected_drugs,
        'is_filtered': bool(select_all or drug_ids),
        'today_jalali': today_jalali, # اضافه شد
    })



@login_required
def drug_print_report_view(request):
    # فراخوانی تابع کمکی برای دریافت دقیق همان داده‌ها
    transactions, selected_drugs, is_filtered = get_drug_transactions_data(request)
    
    context = {
        'report_title': "گزارش چاپی کارتکس داروخانه",
        'transactions': transactions,
        'selected_drugs': selected_drugs,
        'is_filtered': is_filtered,
        'start_date': request.GET.get('start_date'),
        'end_date': request.GET.get('end_date'),
        'today_jalali': JalaliDate.today().strftime('%Y/%m/%d'),
    }
    return render(request, 'reports/drug_transaction_print.html', context)



def drug_summary_report_view(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # تبدیل تاریخ شمسی به میلادی (با استفاده از منطق قبلی خودتان)
    start_dt, end_dt = None, None
    try:
        table = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
        if start_date_str:
            p = [int(x) for x in start_date_str.translate(table).split('/')]
            start_dt = timezone.make_aware(dt_module.datetime.combine(JalaliDate(p[0],p[1],p[2]).to_gregorian(), dt_module.time.min))
        if end_date_str:
            p = [int(x) for x in end_date_str.translate(table).split('/')]
            end_dt = timezone.make_aware(dt_module.datetime.combine(JalaliDate(p[0],p[1],p[2]).to_gregorian(), dt_module.time.max))
    except: pass

    summary_data = []
    # اگر بازه زمانی انتخاب نشده باشد، گزارشی چاپ نمی‌کنیم یا بازه پیش‌فرض می‌گذاریم
    if start_dt and end_dt:
        drugs = Drug.objects.all()
        for drug in drugs:
            # ۱. موجودی اول دوره (تمام ورودی‌ها منهای خروجی‌های قبل از start_dt)
            p_before = PurchaseInvoiceItem.objects.filter(drug=drug, invoice__invoice_date__lt=start_dt).aggregate(s=Sum('quantity'))['s'] or 0
            v_before = VisitItem.objects.filter(drug=drug, visit__visit_date__lt=start_dt).aggregate(s=Sum('quantity'))['s'] or 0
            opening_balance = p_before - v_before

            # ۲. مجموع ورود در بازه
            total_in = PurchaseInvoiceItem.objects.filter(drug=drug, invoice__invoice_date__range=(start_dt, end_dt)).aggregate(s=Sum('quantity'))['s'] or 0
            
            # ۳. مجموع خروج در بازه
            total_out = VisitItem.objects.filter(drug=drug, visit__visit_date__range=(start_dt, end_dt)).aggregate(s=Sum('quantity'))['s'] or 0

            # ۴. مانده پایان دوره
            closing_balance = opening_balance + total_in - total_out

            # فقط داروهایی که گردش داشته‌اند یا موجودی دارند را لیست می‌کنیم
            if opening_balance != 0 or total_in != 0 or total_out != 0:
                summary_data.append({
                    'drug_name': drug.name,
                    'opening': opening_balance,
                    'in': total_in,
                    'out': total_out,
                    'closing': closing_balance,
                })

    context = {
        'report_title': "گزارش خلاصه گردش موجودی (تجمعی)",
        'summary_data': summary_data,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'today_jalali': JalaliDate.today().strftime('%Y/%m/%d'),
    }
    
    # اگر پارامتر print در URL بود، قالب چاپ را رندر کن
    if request.GET.get('print') == '1':
        return render(request, 'reports/drug_summary_print.html', context)
    
    return render(request, 'reports/drug_summary_report.html', context)    